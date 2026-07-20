from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib import messages
from django.db.models import Q
from django.views.generic import ListView, DetailView, CreateView, UpdateView
from django.urls import reverse_lazy
from decimal import Decimal, InvalidOperation
from datetime import date

from shared.mixins import PermissionRequiredRedirectMixin
from .models import Prestamo, PrestamoDetalle, Empleado, TipoPrestamo
from .forms import PrestamoForm


class PrestamoListView(LoginRequiredMixin, PermissionRequiredRedirectMixin, ListView):
    model = Prestamo
    template_name = 'RRHumanos/prestamo_list.html'
    context_object_name = 'prestamos'
    permission_required = 'RRHumanos.view_prestamo'
    permission_redirect_url = '/'

    def get_queryset(self):
        queryset = super().get_queryset().select_related('empleado', 'tipo_prestamo')
        search_query = self.request.GET.get('search', '')
        if search_query:
            queryset = queryset.filter(
                Q(empleado__nombres__icontains=search_query) |
                Q(tipo_prestamo__descripcion__icontains=search_query)
            )
        return queryset


class PrestamoDetailView(LoginRequiredMixin, PermissionRequiredRedirectMixin, DetailView):
    model = Prestamo
    template_name = 'RRHumanos/prestamo_detail.html'
    context_object_name = 'prestamo'
    permission_required = 'RRHumanos.view_prestamo'
    permission_redirect_url = '/'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['detalles'] = self.object.detalles.all()
        context['hoy'] = date.today()
        # Progreso amortizado
        total_pagar = self.object.monto_pagar
        if total_pagar > 0:
            pagado = total_pagar - self.object.saldo
            progreso = int((pagado / total_pagar) * 100)
        else:
            progreso = 0
        context['progreso'] = progreso
        return context


class PrestamoCreateView(LoginRequiredMixin, PermissionRequiredRedirectMixin, CreateView):
    model = Prestamo
    form_class = PrestamoForm
    template_name = 'RRHumanos/prestamo_form.html'
    success_url = reverse_lazy('RRHumanos:prestamo_list')
    permission_required = 'RRHumanos.add_prestamo'
    permission_redirect_url = '/'

    def form_valid(self, form):
        messages.success(self.request, 'El préstamo ha sido registrado y sus cuotas se han generado correctamente.')
        return super().form_valid(form)


class PrestamoUpdateView(LoginRequiredMixin, PermissionRequiredRedirectMixin, UpdateView):
    model = Prestamo
    form_class = PrestamoForm
    template_name = 'RRHumanos/prestamo_form.html'
    permission_required = 'RRHumanos.change_prestamo'
    permission_redirect_url = '/'

    def get_success_url(self):
        return reverse_lazy('RRHumanos:prestamo_detail', kwargs={'pk': self.object.pk})

    def form_valid(self, form):
        if self.get_object().estado != 'PEND':
            messages.error(self.request, 'Solo se pueden editar préstamos en estado Pendiente.')
            return self.form_invalid(form)
        messages.success(self.request, 'El préstamo ha sido actualizado correctamente.')
        return super().form_valid(form)


@login_required
@permission_required('RRHumanos.change_prestamo', login_url='/')
def prestamo_anular(request, pk):
    from django.db import transaction
    with transaction.atomic():
        # Bloquear fila del préstamo
        prestamo = get_object_or_404(Prestamo.objects.select_for_update(), pk=pk)
        
        if prestamo.estado == 'ANU':
            messages.warning(request, 'El préstamo ya está anulado.')
        elif prestamo.estado == 'PAG':
            messages.error(request, 'No se puede anular un préstamo que ya ha sido pagado.')
        elif prestamo.saldo < prestamo.monto_pagar:
            messages.error(request, 'No se puede anular un préstamo con abonos registrados.')
        else:
            prestamo.estado = 'ANU'
            # Resetear saldos a cero
            for detalle in prestamo.detalles.all():
                detalle.saldo_cuota = 0
                detalle.save()
            prestamo.saldo = Decimal('0.00')
            prestamo.save()
            messages.success(request, 'El préstamo ha sido anulado correctamente.')
            
    return redirect('RRHumanos:prestamo_detail', pk=pk)


@login_required
@permission_required('RRHumanos.change_prestamo', login_url='/')
def registrar_pago_cuota(request, cuota_id):
    from django.db import transaction
    
    cuota = get_object_or_404(PrestamoDetalle, pk=cuota_id)
    prestamo = cuota.prestamo

    if request.method == 'POST':
        monto_abono_str = request.POST.get('monto_abono')
        try:
            monto_abono = Decimal(monto_abono_str)
        except (TypeError, ValueError, InvalidOperation):
            messages.error(request, 'El monto del abono no es válido.')
            return redirect('RRHumanos:prestamo_detail', pk=prestamo.pk)

        if monto_abono <= 0:
            messages.error(request, 'El monto del abono debe ser mayor a 0.')
            return redirect('RRHumanos:prestamo_detail', pk=prestamo.pk)

        if monto_abono > cuota.saldo_cuota:
            messages.error(request, f'El monto del abono (${monto_abono:.2f}) no puede ser mayor al saldo de la cuota (${cuota.saldo_cuota:.2f}).')
            return redirect('RRHumanos:prestamo_detail', pk=prestamo.pk)

        with transaction.atomic():
            # Control de concurrencia: select_for_update() en el maestro
            prestamo_locked = Prestamo.objects.select_for_update().get(pk=prestamo.pk)
            
            # Recargar cuota bajo bloqueo
            cuota_locked = PrestamoDetalle.objects.select_for_update().get(pk=cuota_id)
            
            if prestamo_locked.estado == 'ANU':
                messages.error(request, 'No se pueden registrar pagos en un préstamo anulado.')
                return redirect('RRHumanos:prestamo_detail', pk=prestamo.pk)

            # Actualizar saldo de la cuota
            cuota_locked.saldo_cuota -= monto_abono
            cuota_locked.save()

            # Actualizar saldo del maestro
            prestamo_locked.saldo -= monto_abono
            if prestamo_locked.saldo <= 0:
                prestamo_locked.saldo = Decimal('0.00')
                prestamo_locked.estado = 'PAG'
            prestamo_locked.save()

            messages.success(request, f'Se registró el abono de ${monto_abono:.2f} a la cuota #{cuota_locked.numero_cuota}.')
            
    return redirect('RRHumanos:prestamo_detail', pk=prestamo.pk)


@login_required
@permission_required('RRHumanos.view_prestamo', login_url='/')
def exportar_prestamos_excel(request):
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from django.http import HttpResponse
    import io

    # Lógica de filtros similar a la búsqueda
    queryset = Prestamo.objects.select_related('empleado', 'tipo_prestamo').all()
    search_query = request.GET.get('search', request.GET.get('q', ''))
    if search_query:
        queryset = queryset.filter(
            Q(empleado__nombres__icontains=search_query) |
            Q(tipo_prestamo__descripcion__icontains=search_query)
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte Préstamos"

    # Estilos de Excel
    font_bold = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    fill_header = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    align_center = Alignment(horizontal='center', vertical='center')

    # Encabezados
    headers = [
        "ID", "Fecha", "Empleado", "Tipo Préstamo", 
        "Monto ($)", "Interés ($)", "Total a Pagar ($)", "Saldo Pendiente ($)", "Estado"
    ]
    ws.append(headers)

    # Aplicar estilos a la cabecera
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = font_bold
        cell.fill = fill_header
        cell.alignment = align_center

    # Datos
    for item in queryset:
        row = [
            item.id,
            item.fecha_prestamo.strftime("%d/%m/%Y"),
            item.empleado.nombres,
            item.tipo_prestamo.descripcion,
            float(item.monto),
            float(item.interes),
            float(item.monto_pagar),
            float(item.saldo),
            item.get_estado_display()
        ]
        ws.append(row)

    # Ajustar ancho de columnas
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    buffer = io.BytesIO()
    wb.save(buffer)
    xlsx_bytes = buffer.getvalue()
    buffer.close()

    response = HttpResponse(
        xlsx_bytes, 
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="reporte_prestamos.xlsx"'
    return response


@login_required
@permission_required('RRHumanos.view_prestamo', login_url='/')
def exportar_cronograma_pdf(request, prestamo_id):
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from django.http import HttpResponse
    import io

    prestamo = get_object_or_404(Prestamo, pk=prestamo_id)
    detalles = prestamo.detalles.all().order_by('numero_cuota')

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter,
        topMargin=2*cm, bottomMargin=2*cm,
        leftMargin=2*cm, rightMargin=2*cm)

    styles = getSampleStyleSheet()
    story = []

    title_style = ParagraphStyle('title', parent=styles['Normal'],
        fontSize=18, fontName='Helvetica-Bold',
        textColor=colors.HexColor('#1e293b'), spaceAfter=4)
    subtitle_style = ParagraphStyle('subtitle', parent=styles['Normal'],
        fontSize=11, fontName='Helvetica-Bold',
        textColor=colors.HexColor('#4e54c8'), spaceAfter=15)
    normal = ParagraphStyle('normal', parent=styles['Normal'],
        fontSize=9, textColor=colors.HexColor('#334155'), spaceAfter=2)
    bold = ParagraphStyle('bold', parent=styles['Normal'],
        fontSize=9, fontName='Helvetica-Bold',
        textColor=colors.HexColor('#334155'), spaceAfter=2)

    # 1. Cabecera institucional
    story.append(Paragraph('DEPARTAMENTO DE TALENTO HUMANO', title_style))
    story.append(Paragraph('CRONOGRAMA Y COMPROMISO DE PAGO DE PRÉSTAMO', subtitle_style))
    story.append(Paragraph(f'<b>Fecha de Emisión:</b> {date.today().strftime("%d/%m/%Y")}', normal))
    story.append(Spacer(1, 0.4*cm))
    story.append(HRFlowable(width='100%', thickness=1, color=colors.HexColor('#cbd5e1')))
    story.append(Spacer(1, 0.4*cm))

    # 2. Información del préstamo
    info_data = [
        [Paragraph('<b>Empleado:</b>', bold), Paragraph(prestamo.empleado.nombres, normal),
         Paragraph('<b>Monto Solicitado:</b>', bold), Paragraph(f'${prestamo.monto:.2f}', normal)],
        [Paragraph('<b>Sueldo Mensual:</b>', bold), Paragraph(f'${prestamo.empleado.sueldo:.2f}', normal),
         Paragraph('<b>Interés Total:</b>', bold), Paragraph(f'${prestamo.interes:.2f} ({prestamo.tipo_prestamo.tasa_interes}%)', normal)],
        [Paragraph('<b>Tipo Préstamo:</b>', bold), Paragraph(prestamo.tipo_prestamo.descripcion, normal),
         Paragraph('<b>Monto Total a Pagar:</b>', bold), Paragraph(f'${prestamo.monto_pagar:.2f}', normal)],
        [Paragraph('<b>N° de Cuotas:</b>', bold), Paragraph(str(prestamo.numero_cuotas), normal),
         Paragraph('<b>Saldo Pendiente:</b>', bold), Paragraph(f'${prestamo.saldo:.2f}', normal)],
    ]
    info_table = Table(info_data, colWidths=[3.5*cm, 5*cm, 4*cm, 5*cm])
    info_table.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING', (0,0), (-1,-1), 4),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
    ]))
    story.append(info_table)
    story.append(Spacer(1, 0.5*cm))
    story.append(HRFlowable(width='100%', thickness=1, color=colors.HexColor('#cbd5e1')))
    story.append(Spacer(1, 0.5*cm))

    # 3. Tabla de cuotas (Amortización)
    th_style = ParagraphStyle('th', parent=styles['Normal'], fontSize=9, fontName='Helvetica-Bold', textColor=colors.white)
    table_data = [
        [Paragraph('Cuota N°', th_style), 
         Paragraph('Fecha Vencimiento', th_style), 
         Paragraph('Valor Cuota', th_style), 
         Paragraph('Saldo Cuota', th_style), 
         Paragraph('Estado', th_style)]
    ]
    
    for c in detalles:
        estado_cuota = 'Pagada' if c.saldo_cuota == 0 else 'Pendiente'
        table_data.append([
            Paragraph(f'Cuota {c.numero_cuota}', normal),
            Paragraph(c.fecha_vencimiento.strftime("%d/%m/%Y"), normal),
            Paragraph(f'${c.valor_cuota:.2f}', normal),
            Paragraph(f'${c.saldo_cuota:.2f}', normal),
            Paragraph(estado_cuota, normal),
        ])

    cuotas_table = Table(table_data, colWidths=[3*cm, 4.5*cm, 3.5*cm, 3.5*cm, 3*cm])
    cuotas_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1e293b')),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#e2e8f0')),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f8fafc')])
    ]))
    story.append(cuotas_table)
    story.append(Spacer(1, 1.5*cm))

    # 4. Firmas de compromiso
    firma_style = ParagraphStyle('firma', parent=styles['Normal'], fontSize=9, alignment=1)
    firmas_data = [
        ['', ''],
        [Paragraph('_______________________________<br/><b>Firma del Empleado</b><br/>C.I. / RUC', firma_style),
         Paragraph('_______________________________<br/><b>Firma de Talento Humano</b><br/>Autorizado por Empresa', firma_style)]
    ]
    firmas_table = Table(firmas_data, colWidths=[8.5*cm, 9*cm])
    firmas_table.setStyle(TableStyle([
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'BOTTOM'),
        ('TOPPADDING', (0,0), (-1,-1), 10),
    ]))
    story.append(firmas_table)

    doc.build(story)
    pdf = buffer.getvalue()
    buffer.close()

    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="cronograma_prestamo_{prestamo.id}.pdf"'
    return response


@login_required
def api_empleado_sueldo(request, pk):
    from django.http import JsonResponse
    empleado = get_object_or_404(Empleado, pk=pk)
    return JsonResponse({'sueldo': float(empleado.sueldo)})


@login_required
def api_tipo_prestamo_tasa(request, pk):
    from django.http import JsonResponse
    tipo = get_object_or_404(TipoPrestamo, pk=pk)
    return JsonResponse({'tasa_interes': tipo.tasa_interes})


