import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from django.http import HttpResponse
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm


class ExportMixin:
    """
    Mixin genérico para exportar listados a Excel y PDF.
    """

    export_filename = 'listado'
    export_title = 'Listado'
    export_headers = []

    def get_export_rows(self, queryset):
        raise NotImplementedError('Debes implementar get_export_rows()')

    def export_to_excel(self, queryset):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = self.export_title[:31]

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(self.export_headers))
        title_cell = ws.cell(row=1, column=1, value=self.export_title)
        title_cell.font = Font(size=14, bold=True)
        title_cell.alignment = Alignment(horizontal='center')

        header_row = 3
        header_fill = PatternFill(start_color='4E54C8', end_color='4E54C8', fill_type='solid')
        for col_idx, header in enumerate(self.export_headers, start=1):
            cell = ws.cell(row=header_row, column=col_idx, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        rows = self.get_export_rows(queryset)
        for row_idx, row_data in enumerate(rows, start=header_row + 1):
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        for col_idx, header in enumerate(self.export_headers, start=1):
            max_length = len(str(header))
            for row_data in rows:
                if col_idx - 1 < len(row_data):
                    length = len(str(row_data[col_idx - 1]))
                    if length > max_length:
                        max_length = length
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max_length + 4

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{self.export_filename}.xlsx"'
        wb.save(response)
        return response

    def export_to_pdf(self, queryset):
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{self.export_filename}.pdf"'

        doc = SimpleDocTemplate(
            response,
            pagesize=landscape(letter),
            topMargin=1.5 * cm,
            bottomMargin=1.5 * cm,
        )
        elements = []
        styles = getSampleStyleSheet()

        elements.append(Paragraph(self.export_title, styles['Title']))
        elements.append(Spacer(1, 0.5 * cm))

        rows = self.get_export_rows(queryset)
        data = [self.export_headers] + [[str(v) for v in row] for row in rows]

        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4E54C8')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F2F2')]),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        elements.append(table)

        doc.build(elements)
        return response